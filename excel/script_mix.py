from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference, series


dict_anos = {}

arquivo1 = load_workbook(filename="gastos3.xlsx")
ws1 = arquivo1["gastos"]
max_linhas = ws1.max_row

for i in range(2, max_linhas + 1):
    dict_anos[ws1["A%s" % i].value] = {"gastos": ws1["B%s" % i].value, "receita": 0}


arquivo2 = load_workbook(filename="receita.xlsx")
ws2 = arquivo2["receita"]
max_linhas = ws2.max_row

for i in range(2, max_linhas + 1):
    dict_anos[ws2["A%s" % i].value]["receita"] = ws2["B%s" % i].value

# novo arquivo

wb = Workbook()
ws = wb.active

ws["A1"] = "Ano"
ws["B1"] = "Gastos"
ws["C1"] = "Receita"

i = 2
for key, value in dict_anos.items():
    ws["A%s" % i] = key
    ws["B%s" % i] = value["gastos"]
    ws["C%s" % i] = value["receita"]
    i = i + 1

chart1 = BarChart()
chart1.type = "col"
chart1.style = 10
chart1.title = "Receita x Gastos"
chart1.y_axis.title = "R$"
chart1.x_axis.title = "Ano"

data = Reference(ws, min_col=2, max_col=3, min_row=1, max_row=i)
anos = Reference(ws, min_col=1, min_row=2, max_row=i)
chart1.add_data(data, titles_from_data=True)
chart1.set_categories(anos)

ws.add_chart(chart1, "A%s" % (i + 2))

wb.save(filename="resultadomix.xlsx")
