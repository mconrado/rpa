from openpyxl import load_workbook

wb = load_workbook(filename="nomes2.xlsx")

planilha = wb["Planilha1"]

for i in range(2, 5):
    print("%s tem %s anos." % (planilha["A%s" % i].value, planilha["B%s" % i].value))

planilha["B5"] = "=SUM(B2:B4)"

planilha["A7"] = "ALUNOS"
planilha.merge_cells("A7:B7")
planilha.unmerge_cells("A7:B7")


wb.save(filename="nomes2.xlsx")
